# Description: This scripts downloads the OFAC sanctions list in XML format, parses the XML file, and extracts the data into an Excel file.
# We have extracted features, address, id, name from the xml and created a excel workbook with 4 sheets.
# Author: Venkatasai Kadamati
# Date: 7-12-2024

import requests
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
import urllib3

# Disable InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# Custom user variables
XML_URL = "https://www.treasury.gov/ofac/downloads/sanctions/1.0/sdn_advanced.xml"
XML_FILE_PATH = "sdn_advanced.xml"
XLSX_FILE_PATH = "output/sdn_output_addres_test_180nulls.xlsx"

NAMESPACE = {
    # "ns": "http://www.un.org/sanctions/1.0"
    "ns": "https://sanctionslistservice.ofac.treas.gov/api/PublicationPreview/exports/ADVANCED_XML"
}

# ! Changelog : deleted mapping dictionaries for feature_type, list_id, sanctions_type


# utility Functions
# util 1 : latest xml downloader
def download_xml(url, file_path):
    """
    Downloads an XML file from the specified URL and saves it to the given file path.

    Args:
        url (str): The URL to download the XML file from.
        file_path (str): The local file path to save the downloaded XML file.

    Returns:
        bool: True if the file was downloaded successfully, False otherwise.
    """
    try:
        with requests.Session() as session:
            response = session.get(url, stream=True, verify=False)
            response.raise_for_status()  # Raise an exception for HTTP errors
            with open(file_path, "wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:  # Filter out keep-alive new chunks
                        file.write(chunk)
        print("XML file downloaded successfully 🔖")
        return True
    except requests.exceptions.RequestException as e:
        print(f"Error downloading XML: {e}")
        return False


# util 2 : xml parser
def parse_xml(file_path):
    """
    Parses an XML file and returns the tree and root elements.

    Args:
        file_path (str): The path to the XML file to be parsed.

    Returns:
        tuple: A tuple containing the parsed XML tree and root element.
    """
    tree = ET.parse(file_path)
    root = tree.getroot()
    return tree, root


# ! Changelog : added new utility mappers for feature_type, list_id, sanctions_type, reliability_value
# util 3 : mappings extractor
def get_mappings(root, ns):
    """
    Extracts and returns mappings for country, document type, and feature type from the XML root.

    Args:
        root (Element): The root element of the parsed XML tree.
        ns (dict): The namespace dictionary for XML parsing.

    Returns:
        tuple: A tuple containing three dictionaries for country, document type, and feature type mappings.
    """
    country_mapping = {}
    country_values = root.find(".//ns:CountryValues", ns)
    for country in country_values.findall(".//ns:Country", ns):
        country_id = country.attrib["ID"]
        country_name = country.text
        country_mapping[country_id] = country_name

    doc_type_mapping = {}
    doc_type_values = root.find(".//ns:IDRegDocTypeValues", ns)
    for doc_type in doc_type_values.findall(".//ns:IDRegDocType", ns):
        doc_type_id = doc_type.attrib["ID"]
        doc_type_name = doc_type.text
        doc_type_mapping[doc_type_id] = doc_type_name

    list_id_mapping = {}
    list_values = root.find(".//ns:ListValues", ns)
    for list_item in list_values.findall(".//ns:List", ns):
        list_id = list_item.attrib["ID"]
        list_name = list_item.text
        list_id_mapping[list_id] = list_name

    sanctions_type_mapping = {}
    sanctions_type_values = root.find(".//ns:SanctionsTypeValues", ns)
    for sanctions_type in sanctions_type_values.findall(".//ns:SanctionsType", ns):
        sanctions_type_id = sanctions_type.attrib["ID"]
        sanctions_type_name = sanctions_type.text
        sanctions_type_mapping[sanctions_type_id] = sanctions_type_name

    feature_type_mapping = {}
    feature_type_values = root.find(".//ns:FeatureTypeValues", ns)
    for feature_type in feature_type_values.findall(".//ns:FeatureType", ns):
        feature_type_id = feature_type.attrib["ID"]
        feature_type_name = feature_type.text
        feature_type_mapping[feature_type_id] = feature_type_name

    reliability_mapping = {}
    reliability_values = root.find(".//ns:ReliabilityValues", ns)
    for reliability in reliability_values.findall(".//ns:Reliability", ns):
        reliability_id = reliability.attrib["ID"]
        reliability_name = reliability.text
        reliability_mapping[reliability_id] = reliability_name

    detail_reference_mapping = {}
    detail_reference_values = root.find(".//ns:DetailReferenceValues", ns)
    for detail_reference in detail_reference_values.findall(
        ".//ns:DetailReference", ns
    ):
        detail_reference_id = detail_reference.attrib["ID"]
        detail_reference_name = detail_reference.text
        detail_reference_mapping[detail_reference_id] = detail_reference_name

    return (
        country_mapping,
        doc_type_mapping,
        list_id_mapping,
        sanctions_type_mapping,
        feature_type_mapping,
        reliability_mapping,
        detail_reference_mapping,
    )


# parser functions
# parser 1 : feature parser
def feature_parser(
    root, ns, feature_type_mapping, reliability_mapping, detail_reference_mapping
):
    """Parses features from the XML root and returns field names and data rows.

    Args:
        root (Element): The root element of the parsed XML tree.
        ns (dict): The namespace dictionary for XML parsing.
        feature_type_mapping (dict): A dictionary mapping feature type IDs to feature type names.
        reliability_mapping (dict): A dictionary mapping reliability IDs to reliability names.
        detail_reference_mapping (dict): A dictionary mapping detail reference IDs to their values.

    Returns:
        tuple: A tuple containing a list of field names and a list of data rows.
    """
    fieldnames = ["FixedRef", "FeatureType", "Value", "ReliabilityValue", "Comment"]
    data_rows = []

    distinct_parties = root.findall(".//ns:DistinctParty", ns)
    for party in distinct_parties:
        fixed_ref = party.attrib["FixedRef"]
        features = party.findall(".//ns:Feature", ns)
        for feature in features:
            feature_type_id = feature.attrib["FeatureTypeID"]
            feature_type = feature_type_mapping.get(feature_type_id, "")
            feature_version = feature.find(".//ns:FeatureVersion", ns)
            reliability_id = feature_version.attrib.get("ReliabilityID", "")
            reliability_value = reliability_mapping.get(reliability_id, "Unknown")
            comment = (
                feature_version.find(".//ns:Comment", ns).text
                if feature_version.find(".//ns:Comment", ns) is not None
                else ""
            )

            # Get the value based on the feature type
            value = ""
            version_detail = feature_version.find(".//ns:VersionDetail", ns)
            if version_detail is not None:
                detail_type_id = version_detail.attrib.get("DetailTypeID", "")
                if detail_type_id == "1431":  # LOOKUP
                    detail_reference_id = version_detail.attrib.get(
                        "DetailReferenceID", ""
                    )
                    value = detail_reference_mapping.get(detail_reference_id, "")
                elif detail_type_id == "1432":  # TEXT
                    value = version_detail.text
                elif detail_type_id == "1433":  # COUNTRY
                    value = version_detail.attrib.get("CountryID", "")
                elif detail_type_id == "1430":  # DATE
                    date_period = version_detail.find(".//ns:DatePeriod", ns)
                    if date_period is not None:
                        start = date_period.find(".//ns:Start", ns)
                        end = date_period.find(".//ns:End", ns)
                        if start is not None and end is not None:
                            start_date = f"{start.find('.//ns:Year', ns).text}-{start.find('.//ns:Month', ns).text}-{start.find('.//ns:Day', ns).text}"
                            end_date = f"{end.find('.//ns:Year', ns).text}-{end.find('.//ns:Month', ns).text}-{end.find('.//ns:Day', ns).text}"
                            value = f"{start_date} to {end_date}"
                        elif start is not None:
                            start_date = f"{start.find('.//ns:Year', ns).text}-{start.find('.//ns:Month', ns).text}-{start.find('.//ns:Day', ns).text}"
                            value = f"From {start_date}"
                        elif end is not None:
                            end_date = f"{end.find('.//ns:Year', ns).text}-{end.find('.//ns:Month', ns).text}-{end.find('.//ns:Day', ns).text}"
                            value = f"Until {end_date}"

            # Special handling for Location
            if feature_type == "Location":
                location = feature_version.find(".//ns:VersionLocation", ns)
                if location is not None:
                    value = location.attrib.get("LocationID", "")

            data = {
                "FixedRef": fixed_ref,
                "FeatureType": feature_type,
                "Value": value,
                "ReliabilityValue": reliability_value,
                "Comment": comment,
            }
            data_rows.append(data)

    return fieldnames, data_rows


# parser 2 : id parser
def id_parser(root, ns, country_mapping, doc_type_mapping):
    """Parses ID registration documents from the XML root and returns field names and data rows."""
    fieldnames = [
        "FixedRef",
        "Document_Type_ID",
        "Document_Type_Name",
        "Issued_By",
        "Issuing_Country_ID",
        "Issuing_Country_Name",
        "Issue_Date",
        "Expiration_Date",
        "Value",
    ]
    data_rows = []

    for idregdocument in root.findall(".//ns:IDRegDocument", ns):
        identity_id = idregdocument.attrib["IdentityID"]
        distinct_party = root.find(
            f".//ns:DistinctParty/ns:Profile/ns:Identity[@ID='{identity_id}']", ns
        )
        if distinct_party is not None:
            fixed_ref = distinct_party.attrib["FixedRef"]
            document_type_id = idregdocument.attrib["IDRegDocTypeID"]
            document_type_name = doc_type_mapping.get(
                document_type_id, "Unknown Document Type"
            )
            issued_by = (
                idregdocument.find(".//ns:IssuingAuthority", ns).text
                if idregdocument.find(".//ns:IssuingAuthority", ns) is not None
                else ""
            )
            issued_by_country_id = idregdocument.attrib.get("IssuedBy-CountryID", "")
            issued_by_country_name = country_mapping.get(
                issued_by_country_id, "Unknown Country"
            )
            value = (
                idregdocument.find(".//ns:IDRegistrationNo", ns).text
                if idregdocument.find(".//ns:IDRegistrationNo", ns) is not None
                else ""
            )
            issue_date = ""
            expiration_date = ""

            for documentdate in idregdocument.findall(".//ns:DocumentDate", ns):
                idregdocdatetypeid = documentdate.attrib["IDRegDocDateTypeID"]
                dateperiod = documentdate.find(".//ns:DatePeriod", ns)
                if dateperiod is not None:
                    start = dateperiod.find(".//ns:Start", ns)
                    if start is not None:
                        start_year = (
                            start.find(".//ns:Year", ns).text
                            if start.find(".//ns:Year", ns) is not None
                            else ""
                        )
                        start_month = (
                            start.find(".//ns:Month", ns).text
                            if start.find(".//ns:Month", ns) is not None
                            else ""
                        )
                        start_day = (
                            start.find(".//ns:Day", ns).text
                            if start.find(".//ns:Day", ns) is not None
                            else ""
                        )
                        issue_date = f"{start_year}-{start_month}-{start_day}"
                    end = dateperiod.find(".//ns:End", ns)
                    if end is not None:
                        end_year = (
                            end.find(".//ns:Year", ns).text
                            if end.find(".//ns:Year", ns) is not None
                            else ""
                        )
                        end_month = (
                            end.find(".//ns:Month", ns).text
                            if end.find(".//ns:Month", ns) is not None
                            else ""
                        )
                        end_day = (
                            end.find(".//ns:Day", ns).text
                            if end.find(".//ns:Day", ns) is not None
                            else ""
                        )
                        expiration_date = f"{end_year}-{end_month}-{end_day}"
                if idregdocdatetypeid == "1480":
                    issue_date = issue_date
                elif idregdocdatetypeid == "1481":
                    expiration_date = expiration_date

            data = {
                "FixedRef": fixed_ref,
                "Document_Type_ID": document_type_id,
                "Document_Type_Name": document_type_name,
                "Issued_By": issued_by,
                "Issuing_Country_ID": issued_by_country_id,
                "Issuing_Country_Name": issued_by_country_name,
                "Issue_Date": issue_date,
                "Expiration_Date": expiration_date,
                "Value": value,
            }
            data_rows.append(data)
    return fieldnames, data_rows


def address_parser(root, ns, country_mapping):
    """Parses addresses from the XML root and returns field names and data rows."""
    fieldnames = [
        "ID",
        "FixedRef",
        "AreaCodeID",
        "Country",
        "CountryRelevanceID",
        "FeatureVersionID",
        "Unknown",
        "Region",
        "Address 1",
        "Address 2",
        "Address 3",
        "City",
        "State/ Province",
        "Postal Code",
        "Script Type",
    ]
    data_rows = []

    # Create a mapping from IdentityID to FixedRef
    identity_to_fixed_ref = {}
    for party in root.findall(".//ns:DistinctParty", ns):
        fixed_ref = party.attrib["FixedRef"]
        for profile in party.findall(".//ns:Profile", ns):
            for identity in profile.findall(".//ns:Identity", ns):
                identity_id = identity.attrib["ID"]
                identity_to_fixed_ref[identity_id] = fixed_ref

    # Track the first occurrence of each ID to set the Script Type to "Latin"
    first_occurrence = set()

    # Process each Location and write data to CSV
    locations = root.findall(".//ns:Location", ns)
    for location in locations:
        location_id = location.attrib["ID"]
        area_code_id = location.find(".//ns:LocationAreaCode", ns)
        area_code_id = (
            area_code_id.attrib["AreaCodeID"] if area_code_id is not None else ""
        )

        country = location.find(".//ns:LocationCountry", ns)
        country_id = country.attrib["CountryID"] if country is not None else ""

        # Added condition to set country to "undetermined" for area code 11291
        if area_code_id == "11291" and not country_id:
            country_name = "undetermined"
        else:
            country_name = country_mapping.get(country_id, "")

        # Retrieve FixedRef using IdentityID
        id_reg_document_ref = location.find(".//ns:IDRegDocumentReference", ns)
        if id_reg_document_ref is not None:
            identity_id = id_reg_document_ref.attrib["IDRegDocumentID"]
            fixed_ref = identity_to_fixed_ref.get(identity_id, "")
        else:
            fixed_ref = ""

        # Initialize data dictionary
        data = {
            "ID": location_id,
            "FixedRef": fixed_ref,
            "AreaCodeID": area_code_id,
            "Country": country_name,
            "FeatureVersionID": "",
            "Unknown": "",
            "Region": "",
            "Address 1": "",
            "Address 2": "",
            "Address 3": "",
            "City": "",
            "State/ Province": "",
            "Postal Code": "",
            "Script Type": "",
        }

        # Collect non-Latin script values
        non_latin_data = {
            "Chinese Simplified": {
                "Unknown": "",
                "Region": "",
                "Address 1": "",
                "Address 2": "",
                "Address 3": "",
                "City": "",
                "State/ Province": "",
                "Postal Code": "",
            },
            "Chinese Traditional": {
                "Unknown": "",
                "Region": "",
                "Address 1": "",
                "Address 2": "",
                "Address 3": "",
                "City": "",
                "State/ Province": "",
                "Postal Code": "",
            },
            "Cyrillic": {
                "Unknown": "",
                "Region": "",
                "Address 1": "",
                "Address 2": "",
                "Address 3": "",
                "City": "",
                "State/ Province": "",
                "Postal Code": "",
            },
            "Arabic": {
                "Unknown": "",
                "Region": "",
                "Address 1": "",
                "Address 2": "",
                "Address 3": "",
                "City": "",
                "State/ Province": "",
                "Postal Code": "",
            },
            "Japanese": {
                "Unknown": "",
                "Region": "",
                "Address 1": "",
                "Address 2": "",
                "Address 3": "",
                "City": "",
                "State/ Province": "",
                "Postal Code": "",
            },
        }

        for part in location.findall(".//ns:LocationPart", ns):
            part_type_id = part.attrib["LocPartTypeID"]
            for part_value in part.findall(".//ns:LocationPartValue", ns):
                value = (
                    part_value.find(".//ns:Value", ns).text
                    if part_value.find(".//ns:Value", ns) is not None
                    else ""
                )
                comment = (
                    part_value.find(".//ns:Comment", ns).text
                    if part_value.find(".//ns:Comment", ns) is not None
                    else ""
                )

                if not comment:
                    if part_type_id == "1":
                        data["Unknown"] = value
                    elif part_type_id == "1450":
                        data["Region"] = value
                    elif part_type_id == "1451":
                        data["Address 1"] = value
                    elif part_type_id == "1452":
                        data["Address 2"] = value
                    elif part_type_id == "1453":
                        data["Address 3"] = value
                    elif part_type_id == "1454":
                        data["City"] = value
                    elif part_type_id == "1455":
                        data["State/ Province"] = value
                    elif part_type_id == "1456":
                        data["Postal Code"] = value
                else:
                    if comment not in non_latin_data:
                        non_latin_data[comment] = {
                            "Unknown": "",
                            "Region": "",
                            "Address 1": "",
                            "Address 2": "",
                            "Address 3": "",
                            "City": "",
                            "State/ Province": "",
                            "Postal Code": "",
                        }

                    if part_type_id == "1":
                        non_latin_data[comment]["Unknown"] = value
                    elif part_type_id == "1450":
                        non_latin_data[comment]["Region"] = value
                    elif part_type_id == "1451":
                        non_latin_data[comment]["Address 1"] = value
                    elif part_type_id == "1452":
                        non_latin_data[comment]["Address 2"] = value
                    elif part_type_id == "1453":
                        non_latin_data[comment]["Address 3"] = value
                    elif part_type_id == "1454":
                        non_latin_data[comment]["City"] = value
                    elif part_type_id == "1455":
                        non_latin_data[comment]["State/ Province"] = value
                    elif part_type_id == "1456":
                        non_latin_data[comment]["Postal Code"] = value

        # Set Script Type to "Latin" for the first occurrence of each ID
        if data["ID"] not in first_occurrence:
            data["Script Type"] = "Latin"
            first_occurrence.add(data["ID"])

        # Add the Latin script values to data_rows
        data_rows.append(data)

        # Add the non-Latin script values to data_rows
        for script_type, values in non_latin_data.items():
            if (
                values["Unknown"]
                or values["Region"]
                or values["Address 1"]
                or values["Address 2"]
                or values["Address 3"]
                or values["City"]
                or values["State/ Province"]
                or values["Postal Code"]
            ):
                non_latin_row = data.copy()
                non_latin_row["Unknown"] = values["Unknown"]
                non_latin_row["Region"] = values["Region"]
                non_latin_row["Address 1"] = values["Address 1"]
                non_latin_row["Address 2"] = values["Address 2"]
                non_latin_row["Address 3"] = values["Address 3"]
                non_latin_row["City"] = values["City"]
                non_latin_row["State/ Province"] = values["State/ Province"]
                non_latin_row["Postal Code"] = values["Postal Code"]
                non_latin_row["Script Type"] = script_type
                data_rows.append(non_latin_row)

    return fieldnames, data_rows


def name_parser(
    root, ns, script_values, party_subtype_values, alias_type_values, name_part_type_map
):
    def format_name(name_parts):
        name_dict = {
            "Last Name": "",
            "First Name": "",
            "Middle Name": "",
            "Maiden Name": "",
            "Patronymic": "",
            "Matronymic": "",
            "Nickname": "",
            "Entity Name": "",
            "Aircraft Name": "",
            "Vessel Name": "",
        }
        for part in name_parts:
            name_part_group_id = part.attrib["NamePartGroupID"]
            name_part_value = part.text.strip('"')
            name_part_type_id = name_part_type_map.get(name_part_group_id, None)
            if name_part_type_id == "1520":
                name_dict["Last Name"] = name_part_value
            elif name_part_type_id == "1521":
                name_dict["First Name"] = name_part_value
            elif name_part_type_id == "1522":
                name_dict["Middle Name"] = name_part_value
            elif name_part_type_id == "1523":
                name_dict["Maiden Name"] = name_part_value
            elif name_part_type_id == "91708":
                name_dict["Patronymic"] = name_part_value
            elif name_part_type_id == "91709":
                name_dict["Matronymic"] = name_part_value
            elif name_part_type_id == "1528":
                name_dict["Nickname"] = name_part_value
            elif name_part_type_id == "1525":
                name_dict["Entity Name"] = name_part_value
            elif name_part_type_id == "1524":
                name_dict["Aircraft Name"] = name_part_value
            elif name_part_type_id == "1526":
                name_dict["Vessel Name"] = name_part_value

        if name_dict["Last Name"] and name_dict["First Name"]:
            return f"{name_dict['Last Name']}, {name_dict['First Name']} {name_dict['Middle Name']} {name_dict['Maiden Name']}".strip()
        elif name_dict["Last Name"]:
            return name_dict["Last Name"]
        elif (
            name_dict["Patronymic"]
            and name_dict["Matronymic"]
            and name_dict["First Name"]
        ):
            return f"{name_dict['Patronymic']} {name_dict['Matronymic']}, {name_dict['First Name']} {name_dict['Middle Name']} {name_dict['Maiden Name']}".strip()
        elif name_dict["Nickname"]:
            return name_dict["Nickname"]
        elif name_dict["Entity Name"]:
            return name_dict["Entity Name"]
        elif name_dict["Aircraft Name"]:
            return name_dict["Aircraft Name"]
        elif name_dict["Vessel Name"]:
            return name_dict["Vessel Name"]
        return ""

    def get_designation(party_subtype_id):
        if party_subtype_id == "1":
            return "Vessel"
        elif party_subtype_id == "2":
            return "Aircraft"
        elif party_subtype_id == "3":
            return "Business"
        elif party_subtype_id == "4":
            return "Individual"
        else:
            return "Unknown"

    fieldnames = [
        "FixedRef",
        "DocumentedNameID",
        "Designation",
        "Primary Entry",
        "Alias Type",
        "Low Quality",
        "Acronym",
        "Script",
        "Name",
    ]
    data_rows = []
    seen_records = set()

    for party in root.findall(".//ns:DistinctParty", ns):
        fixed_ref = party.attrib["FixedRef"]
        for profile in party.findall(".//ns:Profile", ns):
            party_subtype_id = profile.attrib["PartySubTypeID"]
            designation = get_designation(party_subtype_id)

            for identity in profile.findall(".//ns:Identity", ns):
                for alias in identity.findall(".//ns:Alias", ns):
                    alias_type_id = alias.attrib["AliasTypeID"]
                    alias_type = alias_type_values.get(alias_type_id, "Unknown")
                    low_quality = alias.attrib["LowQuality"]
                    primary_entry = alias.attrib["Primary"]
                    for documented_name in alias.findall(".//ns:DocumentedName", ns):
                        documented_name_id = documented_name.attrib["ID"]
                        name_parts = documented_name.findall(
                            ".//ns:DocumentedNamePart/ns:NamePartValue", ns
                        )
                        name = format_name(name_parts)
                        script_id = (
                            name_parts[0].attrib["ScriptID"]
                            if name_parts
                            else "Unknown"
                        )
                        script = script_values.get(script_id, "Unknown")
                        acronym = (
                            name_parts[0].attrib["Acronym"] if name_parts else "false"
                        )
                        record = (
                            fixed_ref,
                            documented_name_id,
                            designation,
                            primary_entry,
                            alias_type,
                            low_quality,
                            acronym,
                            script,
                            name,
                        )
                        if record not in seen_records:
                            data_rows.append(record)
                            seen_records.add(record)
    return fieldnames, data_rows


def sanctions_entries_parser(root, ns, list_id_mapping, sanctions_type_mapping):
    """Parses sanctions entries from the XML root and returns field names and data rows."""
    fieldnames = ["FixedRef", "ListID", "SanctionsTypeID", "SanctionsProgramID"]
    data_rows = []

    for entry in root.findall(".//ns:SanctionsEntry", ns):
        entry_id = entry.attrib.get("ID", "")
        list_id = entry.attrib.get("ListID", "")
        list_name = list_id_mapping.get(list_id, "Unknown List")
        sanctions_measures = entry.findall(".//ns:SanctionsMeasure", ns)
        for measure in sanctions_measures:
            sanctions_type_id = measure.attrib.get("SanctionsTypeID", "")
            sanctions_type = sanctions_type_mapping.get(
                sanctions_type_id, "Unknown Type"
            )
            sanctions_program_id = ""
            comment = measure.find(".//ns:Comment", ns)
            if comment is not None:
                sanctions_program_id = comment.text
            data_rows.append(
                [entry_id, list_name, sanctions_type, sanctions_program_id]
            )
    return fieldnames, data_rows


def main():
    if download_xml(XML_URL, XML_FILE_PATH):
        tree, root = parse_xml(XML_FILE_PATH)
        (
            country_mapping,
            doc_type_mapping,
            list_id_mapping,
            sanctions_type_mapping,
            feature_type_mapping,
            reliability_mapping,
            detail_reference_mapping,
        ) = get_mappings(root, NAMESPACE)

        # Extract reference values for name parser
        script_values = {
            script_elem.attrib["ID"]: script_elem.text
            for script_elem in root.findall(".//ns:ScriptValues/ns:Script", NAMESPACE)
        }
        party_subtype_values = {
            subtype_elem.attrib["ID"]: subtype_elem.text
            for subtype_elem in root.findall(
                ".//ns:PartySubTypeValues/ns:PartySubType", NAMESPACE
            )
        }
        alias_type_values = {
            alias_elem.attrib["ID"]: alias_elem.text
            for alias_elem in root.findall(
                ".//ns:AliasTypeValues/ns:AliasType", NAMESPACE
            )
        }
        name_part_type_map = {
            group.attrib["ID"]: group.attrib["NamePartTypeID"]
            for group in root.findall(
                ".//ns:MasterNamePartGroup/ns:NamePartGroup", NAMESPACE
            )
        }
        # party_type_values = {
        #     type_elem.attrib["ID"]: type_elem.text
        #     for type_elem in root.findall(
        #         ".//ns:PartyTypeValues/ns:PartyType", NAMESPACE
        #     )
        # }

        # Parse features
        feature_fieldnames, feature_data_rows = feature_parser(
            root,
            NAMESPACE,
            feature_type_mapping,
            reliability_mapping,
            detail_reference_mapping,
        )
        # Parse IDs
        id_fieldnames, id_data_rows = id_parser(
            root, NAMESPACE, country_mapping, doc_type_mapping
        )
        # Parse addresses
        address_fieldnames, address_data_rows = address_parser(
            root, NAMESPACE, country_mapping
        )

        # Parse names
        name_fieldnames, name_data_rows = name_parser(
            root,
            NAMESPACE,
            script_values,
            party_subtype_values,
            alias_type_values,
            name_part_type_map,
        )

        # Parse sanctions entries
        sanctions_entries_fieldnames, sanctions_entries_data_rows = (
            sanctions_entries_parser(
                root, NAMESPACE, list_id_mapping, sanctions_type_mapping
            )
        )

        # Create Excel workbook
        wb = Workbook()
        # Add feature data to sheet
        ws_feature = wb.create_sheet("FEATURE")
        ws_feature.append(feature_fieldnames)
        for row in feature_data_rows:
            ws_feature.append([row[field] for field in feature_fieldnames])
        # Add ID data to sheet
        ws_id = wb.create_sheet("ID")
        ws_id.append(id_fieldnames)
        for row in id_data_rows:
            ws_id.append([row[field] for field in id_fieldnames])
        # Add address data to sheet
        ws_address = wb.create_sheet("ADDRESS")
        ws_address.append(address_fieldnames)
        for row in address_data_rows:
            # Ensure 'CountryRelevanceID' is present in the row dictionary
            row.setdefault("CountryRelevanceID", "")
            ws_address.append([row[field] for field in address_fieldnames])
        # Add sanctions entries data to sheet
        ws_sanctions_entries = wb.create_sheet("SANCTIONS_ENTRIES")
        ws_sanctions_entries.append(sanctions_entries_fieldnames)
        for row in sanctions_entries_data_rows:
            ws_sanctions_entries.append(row)
        # Add name data to sheet
        ws_name = wb.create_sheet("NAME")
        ws_name.append(name_fieldnames)
        for row in name_data_rows:
            ws_name.append(row)
        # Remove the default sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        # Save the workbook
        wb.save(XLSX_FILE_PATH)
        print("Excel file created successfully 🎉")


if __name__ == "__main__":
    main()
